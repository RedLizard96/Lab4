import csv
import openpyxl
from datetime import datetime

# Функція для розрахунку віку на основі дати народження
def calculate_age(birth_date):
    today = datetime.today()
    return today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))

# Функція для зчитування даних із CSV файлу
def read_csv(file_name):
    employees = []
    try:
        with open(file_name, newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            next(reader)  # Пропускаємо заголовки
            for row in reader:
                birth_date = datetime.strptime(row[4], '%Y-%m-%d')  # Дата народження у форматі 'рррр-мм-дд'
                age = calculate_age(birth_date)
                employees.append({
                    'Прізвище': row[0],
                    'Ім’я': row[1],
                    'По батькові': row[2],
                    'Дата народження': row[4],
                    'Вік': age
                })
        print("Ok: CSV файл успішно відкрито.")
    except FileNotFoundError:
        print("Помилка: CSV файл не знайдено.")
        return None
    except Exception as e:
        print(f"Помилка: {e}")
        return None
    return employees

# Функція для створення XLSX файлу
def create_xlsx(employees):
    try:
        # Створюємо новий XLSX файл
        wb = openpyxl.Workbook()

        # Створюємо аркуші
        all_sheet = wb.active
        all_sheet.title = "all"
        younger_18_sheet = wb.create_sheet("younger_18")
        category_18_45_sheet = wb.create_sheet("18-45")
        category_45_70_sheet = wb.create_sheet("45-70")
        older_70_sheet = wb.create_sheet("older_70")

        # Заголовки таблиць
        headers = ['№', 'Прізвище', 'Ім’я', 'По батькові', 'Дата народження', 'Вік']
        for sheet in [all_sheet, younger_18_sheet, category_18_45_sheet, category_45_70_sheet, older_70_sheet]:
            sheet.append(headers)

        # Заповнюємо дані на всіх аркушах
        for i, employee in enumerate(employees, 1):
            row = [i, employee['Прізвище'], employee['Ім’я'], employee['По батькові'], employee['Дата народження'], employee['Вік']]
            all_sheet.append(row)

            # Розподіляємо дані за віковими категоріями
            age = employee['Вік']
            if age < 18:
                younger_18_sheet.append(row)
            elif 18 <= age <= 45:
                category_18_45_sheet.append(row)
            elif 45 <= age <= 70:
                category_45_70_sheet.append(row)
            else:
                older_70_sheet.append(row)

        # Зберігаємо файл
        wb.save('employees.xlsx')
        print("Ok: XLSX файл успішно створено.")
    except Exception as e:
        print(f"Помилка при створенні XLSX файлу: {e}")

# Основна функція
def main():
    file_name = 'employees.csv'  # Назва CSV файлу
    employees = read_csv(file_name)

    if employees:
        create_xlsx(employees)

# Запуск основної функції
if __name__ == '__main__':
    main()
